import cv2
import numpy as np
import os
from scipy.spatial.transform import Rotation as Rot
import open3d as o3d
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from image_processing_A import split_image  # Correct import statement
from Open3dToolBox import Open3dToolBox  # Assuming you save the toolbox in a file named open3d_toolbox.py
import time
MIN_MATCH_COUNT = 10

# Preprocess the image and split into ROIs
image_path = 'image.jpg'
split_image(image_path)

# Get the base directory (directory containing image.jpg)
base_dir = os.path.abspath(os.path.dirname(image_path))

# Process each ROI directory
roi_dirs = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]

# Replace K with given Intrinsic Matrix
fx = 2803.248768176667
fy = 2819.8335675313083
cx = 640.0721808721448
cy = 361.4431800035303
K = np.array([[fx,  0., cx],
              [0., fy, cy],
              [ 0.,  0.,  1.]])

counter = 0

def printTf(R, T):
    msg = "Rotation:\n"
    msg += str(R) + "\n"
    msg += "translation\n"
    msg += str(T) + "\n"
    msg += "Eulers roll pitch yaw (deg)\n"
    msg += str(np.flip(Rot.from_matrix(R).as_euler('ZYX', degrees=True))) + "\n"
    return msg

def TtoS(te):
    te = te.flatten()  # Ensure te is a flat array
    s1 = np.zeros((3, 3))
    s1[0, 1] = -te[2]
    s1[0, 2] = te[1]
    s1[1, 0] = te[2]
    s1[1, 2] = -te[0]
    s1[2, 0] = -te[1]
    s1[2, 1] = te[0]
    return s1

# Initialize Open3D visualizer
vis = o3d.visualization.Visualizer()
vis.create_window()
open3dToolBox = Open3dToolBox(vis)

# Create an Excel workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Matrices Output"

# Add headers to the Excel sheet
headers = ['ROI', 'Transformation', 'Rotation', 'Image1', 'Image2', 'Translation_X', 'Translation_Y', 'Translation_Z', 'Euler_X', 'Euler_Y', 'Euler_Z']
ws.append(headers)

# Define shades of blue for ROIs
shades_of_blue = ["E0F7FA", "B2EBF2", "80DEEA", "4DD0E1", "26C6DA"]
color_index = 0

for roi_dir in roi_dirs:
    images_dir = os.path.join(base_dir, roi_dir, 'images')
    if not os.path.exists(images_dir):
        continue

    img_files = [f for f in os.listdir(images_dir) if f.endswith('.jpg')]
    if len(img_files) < 2:
        continue

    img_pair_1 = os.path.join(images_dir, img_files[0])
    img_pair_2 = os.path.join(images_dir, img_files[1])

    counter += 1

    img1 = cv2.imread(img_pair_1)
    img2 = cv2.imread(img_pair_2)

    # SIFT feature matching
    sift = cv2.SIFT_create()
    kp1, des1 = sift.detectAndCompute(img1, None)
    kp2, des2 = sift.detectAndCompute(img2, None)

    flann = cv2.FlannBasedMatcher(dict(algorithm=0, trees=5), dict(checks=50))
    matches = flann.knnMatch(des1, des2, k=2)

    good = []
    for m, n in matches:
        if m.distance < 0.7 * n.distance:
            good.append(m)

    if len(good) > MIN_MATCH_COUNT:
        p1 = np.float32([kp1[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
        p2 = np.float32([kp2[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)

        # Save good matches to a text file
        matches_file_path = os.path.join(base_dir, roi_dir, f'good_matches_{counter}.txt')
        with open(matches_file_path, mode='w') as file:
            for idx, match in enumerate(good):
                point1 = kp1[match.queryIdx].pt
                point2 = kp2[match.trainIdx].pt
                file.write(f"{idx + 1}: Image 1: {point1} <--> Image 2: {point2}\n")

        draw_params = dict(matchColor=(0, 255, 0), singlePointColor=None, flags=2)
        img_siftmatch = cv2.drawMatches(img1, kp1, img2, kp2, good, None, **draw_params)
        cv2.imwrite(os.path.join(base_dir, roi_dir, f'sift_match_{counter}.png'), img_siftmatch)

        # Essential matrix
        E, mask = cv2.findEssentialMat(p1, p2, K, cv2.RANSAC, 0.999, 1.0)

        matchesMask = mask.ravel().tolist()
        draw_params = dict(matchColor=(0, 255, 0), singlePointColor=None, matchesMask=matchesMask, flags=2)
        img_inliermatch = cv2.drawMatches(img1, kp1, img2, kp2, good, None, **draw_params)
        cv2.imwrite(os.path.join(base_dir, roi_dir, f'inlier_match_{counter}.png'), img_inliermatch)
        print("Essential matrix:")
        print(E)

        # Decompose essential matrix
        R1e, R2e, te = cv2.decomposeEssentialMat(E)
        print('R1e' + str(np.flip(Rot.from_matrix(R1e).as_euler('ZYX', degrees=True))))
        print('R2e' + str(np.flip(Rot.from_matrix(R2e).as_euler('ZYX', degrees=True))))
        print('te' + str(te))

        # Define transformations
        E1 = R1e.dot(TtoS(te))  # R1e, te
        E2 = R1e.dot(TtoS(-te)) # R1e, -te
        E3 = R2e.dot(TtoS(te))  # R2e, te
        E4 = R2e.dot(TtoS(-te)) # R2e, -te

        points0, R0, t0, mask0 = cv2.recoverPose(E, p1, p2, mask=mask)
        points1, R1, t1, mask1 = cv2.recoverPose(E1, p1, p2, mask=mask)
        points2, R2, t2, mask2 = cv2.recoverPose(E2, p1, p2, mask=mask)
        points3, R3, t3, mask3 = cv2.recoverPose(E3, p1, p2, mask=mask)
        points4, R4, t4, mask4 = cv2.recoverPose(E4, p1, p2, mask=mask)

        # Prepare poses with descriptive names
        poses = [
            (R0, t0, "Original E", E),
            (R1, t1, "R1,t", E1),
            (R2, t2, "R1,-t", E2),
            (R3, t3, "R2,t", E3),
            (R4, t4, "R2,-t", E4)
        ]

        transformation_file_path = os.path.join(base_dir, roi_dir, f'transformations_{roi_dir}.txt')
        with open(transformation_file_path, mode='w') as file:
            file.write(f"Transformations for ROI: {roi_dir}\n\n")
            file.write(f"Translation Vectors:\n")
            file.write(f"t0:\n{t0}\n")
            file.write(f"t1:\n{t1}\n")
            file.write(f"t2:\n{t2}\n")
            file.write(f"t3:\n{t3}\n")
            file.write(f"t4:\n{t4}\n\n")
            
            file.write(f"Rotation Matrices:\n")
            file.write(f"R0:\n{R0}\n")
            file.write(f"R1:\n{R1}\n")
            file.write(f"R2:\n{R2}\n")
            file.write(f"R3:\n{R3}\n")
            file.write(f"R4:\n{R4}\n\n")

            for i, (R, t, name, E_i) in enumerate(poses):
                T = np.eye(4)
                T[:3, :3] = R
                T[:3, 3] = t.squeeze()
                object_name = f'object_{counter}_{i}'

                # Visualize each transformation separately
                #open3dToolBox.vis.clear_geometries()
                open3dToolBox.displayTriedre([T], [object_name], triedreAxeSize=0.1)
                open3dToolBox.updateRenderer()

                # Save essential matrix and euler angles to text file
                file.write(f"Transformation {name}:\n")
                file.write("Essential Matrix:\n")
                for row in E_i:
                    file.write(" ".join(f"{elem:.6f}" for elem in row) + "\n")
                euler_angles = np.flip(Rot.from_matrix(R).as_euler('ZYX', degrees=True))
                file.write(f"Eulers roll pitch yaw (deg): {euler_angles}\n")
                file.write("\n")

                # Save matrices to Excel file
                row = [roi_dir, name, "Rotation Matrix", img_files[0], img_files[1],
                       f"{t[0, 0]:.2f}", f"{t[1, 0]:.2f}", f"{t[2, 0]:.2f}",
                       f"{euler_angles[0]:.2f}", f"{euler_angles[1]:.2f}", f"{euler_angles[2]:.2f}"]
                ws.append(row)
                fill = PatternFill(start_color=shades_of_blue[color_index % len(shades_of_blue)], end_color=shades_of_blue[color_index % len(shades_of_blue)], fill_type="solid")
                for j in range(len(row)):
                    ws.cell(row=ws.max_row, column=j+1).fill = fill

    color_index += 1

# Save the Excel workbook
excel_file_path = os.path.join(base_dir, 'matrices_output.xlsx')
wb.save(excel_file_path)

while True:
    open3dToolBox.updateRenderer()
    time.sleep(0.1)
# Close the Open3D visualizer window
if vis is not None:
    vis.destroy_window()
